import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Read CSV with proper quoting
data = []
with open('PHASE_1_FEATURES_ESTIMATION.csv', 'r', encoding='utf-8') as f:
    reader = csv.reader(f)
    headers = next(reader)
    for row in reader:
        # Pad row if needed
        while len(row) < len(headers):
            row.append('')
        data.append(row)

# Create Excel workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Phase 1 Features"

# Define styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_alignment = Alignment(horizontal='center', vertical='center')
wrap_alignment = Alignment(wrap_text=True, vertical='top')

# Component colors
component_colors = {
    'Backend': 'E7F3FF',
    'Admin': 'FFF4E6',
    'Frontend': 'E8F5E9',
    'Mobile': 'F3E5F5',
    'Infrastructure': 'FFEBEE',
    'Documentation': 'F1F8E9'
}

# Write headers
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_alignment
    cell.border = border

# Write data
for row_num, row_data in enumerate(data, 2):
    component = row_data[1] if len(row_data) > 1 else ''
    row_color = component_colors.get(component, 'FFFFFF')
    
    for col_num, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_num, value=value)
        cell.border = border
        
        # Apply component-based row color
        if component and row_color:
            cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
        
        # Format Estimated Hours column as number
        if col_num == 6 and value:
            try:
                cell.value = int(value)
                cell.alignment = center_alignment
            except:
                pass
        
        # Wrap text for Description column
        if col_num == 5:
            cell.alignment = wrap_alignment
        
        # Center align Feature ID and Priority
        if col_num == 1 or col_num == 7:
            cell.alignment = center_alignment

# Adjust column widths
column_widths = {
    'A': 12,  # Feature ID
    'B': 15,  # Component
    'C': 20,  # Category
    'D': 40,  # Feature Name
    'E': 60,  # Description
    'F': 15,  # Estimated Hours
    'G': 12,  # Priority
    'H': 15   # Dependencies
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Freeze header row
ws.freeze_panes = 'A2'

# Add summary sheet
summary_ws = wb.create_sheet("Summary")
summary_ws.append(['Component', 'Total Features', 'Total Hours', 'Must-Have Features', 'Should-Have Features'])

component_totals = {}
for row in data:
    component = row[1] if len(row) > 1 else 'Unknown'
    hours = int(row[5]) if len(row) > 5 and row[5].isdigit() else 0
    priority = row[6] if len(row) > 6 else ''
    
    if component not in component_totals:
        component_totals[component] = {'count': 0, 'hours': 0, 'must_have': 0, 'should_have': 0}
    
    component_totals[component]['count'] += 1
    component_totals[component]['hours'] += hours
    if priority == 'Must-Have':
        component_totals[component]['must_have'] += 1
    elif priority == 'Should-Have':
        component_totals[component]['should_have'] += 1

for component, totals in sorted(component_totals.items()):
    summary_ws.append([
        component,
        totals['count'],
        totals['hours'],
        totals['must_have'],
        totals['should_have']
    ])

# Style summary sheet
for row in summary_ws.iter_rows(min_row=1, max_row=1):
    for cell in row:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border

for col in ['A', 'B', 'C', 'D', 'E']:
    summary_ws.column_dimensions[col].width = 20

# Add totals row
total_row = len(component_totals) + 2
summary_ws.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
summary_ws.cell(row=total_row, column=2, value=sum(t['count'] for t in component_totals.values())).font = Font(bold=True)
summary_ws.cell(row=total_row, column=3, value=sum(t['hours'] for t in component_totals.values())).font = Font(bold=True)
summary_ws.cell(row=total_row, column=4, value=sum(t['must_have'] for t in component_totals.values())).font = Font(bold=True)
summary_ws.cell(row=total_row, column=5, value=sum(t['should_have'] for t in component_totals.values())).font = Font(bold=True)

# Save workbook
wb.save('PHASE_1_FEATURES_ESTIMATION.xlsx')
print(f"Excel file created successfully: PHASE_1_FEATURES_ESTIMATION.xlsx")
print(f"Total features: {len(data)}")
print(f"Total hours: {sum(int(row[5]) if len(row) > 5 and row[5].isdigit() else 0 for row in data)}")


